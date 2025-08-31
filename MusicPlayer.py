import json
import os
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import io

import pygame
import requests
from openpyxl import load_workbook
from PIL import Image, ImageTk
from tkinter import messagebox, scrolledtext
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
import pandas as pd
from mutagen.mp3 import MP3
import re
import jieba
from wordcloud import WordCloud
from PIL import Image
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from similarsong import compute_similarity
import random
from songranking import *
# Set request headers
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}

class MainUI:
    paused = False
    def __init__(self, root):
        # Open and resize play icon image
        self.play_icon = Image.open("img/play.png").resize((40, 40), Image.LANCZOS)
        self.play_icon = ImageTk.PhotoImage(self.play_icon)

        # Open and resize stop icon image
        self.stop_icon = Image.open("img/stop.png").resize((40, 40), Image.LANCZOS)
        self.stop_icon = ImageTk.PhotoImage(self.stop_icon)

        # Initialize the main window
        self.root = root
        self.root.title("Music Player")  # Set window title
        self.root.geometry("1300x800")  # Set window size
        self.root.attributes("-alpha", 0.95)  # Set window transparency
        pygame.mixer.init()  # Initialize Pygame mixer
        self.music_list = []  # Initialize music list storage structure
        self.favorite_music_list = []
        self.current_song_index = None  # Index of currently selected song
        self.selected_song_id = None

        self.setup_ui()  # Set up user interface
        #self.load_initial_songs()  # Load some initial song data
        self.playing = False

    def setup_ui(self):
        # Create a Canvas to draw a gradient background
        self.canvas = tk.Canvas(self.root, width=800, height=500)
        self.canvas.pack(fill=tk.BOTH, expand=True)

        # Create a frame on the Canvas
        main_frame = tk.Frame(self.canvas, bg='#F5F5F0')
        main_frame.place(relwidth=1, relheight=1)

        # Create and configure the left button panel
        left_frame = tk.Frame(main_frame, width=250, relief=tk.RAISED, borderwidth=2, bg="#E8EADD")
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)
        left_frame.pack_propagate(False)  # Prevent the frame from automatically adjusting its size to fit its contents

        # Load icons and resize them
        load_icon = Image.open("img/load.png").resize((50, 50), Image.LANCZOS)
        load_icon = ImageTk.PhotoImage(load_icon)

        song_icon = Image.open("img/song.webp").resize((50, 50), Image.LANCZOS)
        song_icon = ImageTk.PhotoImage(song_icon)

        artist_icon = Image.open("img/artist.webp").resize((50, 50), Image.LANCZOS)
        artist_icon = ImageTk.PhotoImage(artist_icon)

        analysis_icon = Image.open("img/analysis.webp").resize((50, 50), Image.LANCZOS)
        analysis_icon = ImageTk.PhotoImage(analysis_icon)

        search_icon = Image.open("img/search.png").resize((30, 30), Image.LANCZOS)
        search_icon = ImageTk.PhotoImage(search_icon)

        download_icon = Image.open("img/download.webp").resize((40, 40), Image.LANCZOS)
        download_icon = ImageTk.PhotoImage(download_icon)

        prev_icon = Image.open("img/prev.png").resize((40, 40), Image.LANCZOS)
        prev_icon = ImageTk.PhotoImage(prev_icon)

        play_icon = Image.open("img/play.png").resize((40, 40), Image.LANCZOS)
        play_icon = ImageTk.PhotoImage(play_icon)

        stop_icon = Image.open("img/stop.png").resize((40, 40), Image.LANCZOS)
        stop_icon = ImageTk.PhotoImage(stop_icon)

        next_icon = Image.open("img/next.png").resize((40, 40), Image.LANCZOS)
        next_icon = ImageTk.PhotoImage(next_icon)

        favorite_icon = Image.open("img/favorite.webp").resize((40, 40), Image.LANCZOS)
        favorite_icon = ImageTk.PhotoImage(favorite_icon)

        collection_icon = Image.open("img/collection.png").resize((50, 50), Image.LANCZOS)
        collection_icon = ImageTk.PhotoImage(collection_icon)

        back_icon = Image.open("img/back.png").resize((50, 50), Image.LANCZOS)
        back_icon = ImageTk.PhotoImage(back_icon)

        cloud_icon = Image.open("img/cloud.png").resize((50, 50), Image.LANCZOS)
        cloud_icon = ImageTk.PhotoImage(cloud_icon)

        lyrics_icon = Image.open("img/lyrics.png").resize((40, 40), Image.LANCZOS)
        lyrics_icon = ImageTk.PhotoImage(lyrics_icon)

        recommend_icon = Image.open("img/recommend.png").resize((40, 40), Image.LANCZOS)
        recommend_icon = ImageTk.PhotoImage(recommend_icon)

        # Load songs button
        load_button = tk.Button(left_frame, image=load_icon,
                                text="Load songs", compound="left", command=self.load_excel,
                                bg="#D2D7BF", fg="#363C2B")
        load_button.image = load_icon
        load_button.pack(fill=tk.X, padx=10, pady=10)

        # Button to display hot songs
        song_button = tk.Button(left_frame, image=song_icon, text="Hot songs", compound="left",
                                command=self.song_function, bg="#B5BD99", fg="#363C2B")
        song_button.image = song_icon
        song_button.pack(fill=tk.X, padx=10, pady=10)

        # Button to display top singer information
        artist_button = tk.Button(left_frame, image=artist_icon, text="Top singers", compound="left",
                                command=self.artist_function, bg="#9AA477", fg="#363C2B")
        artist_button.image = artist_icon
        artist_button.pack(fill=tk.X, padx=10, pady=10)

        # Data analysis button
        analysis_button = tk.Button(left_frame, image=analysis_icon, text="Data analysis", compound="left",
                                    command=self.analysis_function, bg="#808C5C", fg="#363C2B")
        analysis_button.image = analysis_icon
        analysis_button.pack(fill=tk.X, padx=10, pady=10)

        # My favorites button
        myFavorite_button = tk.Button(left_frame, image=collection_icon, text="Favorite songs", compound="left",
                                    command=self.show_favorite_song, bg="#9AA477", fg="#363C2B")
        myFavorite_button.image = collection_icon
        myFavorite_button.pack(fill=tk.X, padx=10, pady=10)

        cloud_button = tk.Button(left_frame, image=cloud_icon,
                                text="Cloud", compound="left",
                                command=lambda: self.generate_lyric_word_cloud(self.selected_song_id),
                                bg="#B5BD99", fg="#363C2B")
        cloud_button.image = cloud_icon
        cloud_button.pack(fill=tk.X, padx=10, pady=10)

        # Back button
        back_button = tk.Button(left_frame, image=back_icon, text="Back", compound="left",
                                command=self.back_function, bg="#D2D7BF", fg="#363C2B")
        back_button.image = back_icon
        back_button.pack(fill=tk.X, padx=10, pady=10)

        # Set the search panel at the top
        top_frame = tk.Frame(main_frame, bg="#F5F5F0")
        top_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)

        # Search box
        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(top_frame, textvariable=self.search_var, bg="#E8EADD")
        self.search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10)

        # Search button
        search_button = tk.Button(top_frame, image=search_icon, command=self.search_music, bg="#B5BD99", fg="#FFFFFF")
        search_button.image = search_icon
        search_button.pack(side=tk.LEFT, padx=10)

        # Create Treeview style
        style = ttk.Style()
        style.configure("Treeview", background="#F5F5F0", fieldbackground="#F5F5F0", foreground="black")

        # Create a scrollbar
        self.scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL)

        # Create a tree view to display the song list
        self.tree = ttk.Treeview(main_frame, columns=('ID', 'Title', 'Album', 'Artist', 'Duration'),
                                show='headings', style="Treeview", yscrollcommand=self.scrollbar.set)
        self.tree.heading('ID', text='ID')
        self.tree.heading('Title', text='Song')
        self.tree.heading('Artist', text='Artist')
        self.tree.heading('Album', text='Album')
        self.tree.heading('Duration', text='Duration')
        self.tree.column('ID', width=0, stretch=False)  # Hide the ID column

        # Bind the scrollbar to the Treeview's y-axis scroll
        self.scrollbar.config(command=self.tree.yview)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Bind the click event
        self.tree.bind('<ButtonRelease-1>', self.on_tree_select)

        # Bottom control button panel
        bottom_frame = tk.Frame(main_frame, height=60, bg="#F5F5F0")
        bottom_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=5, pady=5)

        # Create buttons for download, previous song, play, stop, next song, and favorites
        download_button = tk.Button(bottom_frame, image=download_icon,
                                    text=" download", compound="left", command=self.download_music,
                                    bg="#808C5C", fg="#FFFFFF")
        download_button.image = download_icon
        download_button.pack(side=tk.LEFT, padx=10, pady=10, expand=True)

        prev_button = tk.Button(bottom_frame, image=prev_icon,
                                text=" previous song", compound="left", command=self.prev_song,
                                bg="#808C5C", fg="#FFFFFF")
        prev_button.image = prev_icon
        prev_button.pack(side=tk.LEFT, padx=10, pady=10, expand=True)

        play_button = tk.Button(bottom_frame, image=play_icon,
                                text=" play", compound="left", command=lambda: self.play_music(self.selected_song_id),
                                bg="#808C5C", fg="#FFFFFF")
        play_button.image = play_icon
        play_button.pack(side=tk.LEFT, padx=10, pady=10, expand=True)

        self.current_time_label = tk.Label(self.root, text="current time: 00:00")
        self.current_time_label.pack()

        self.total_time_label = tk.Label(self.root, text="total time: 00:00")
        self.total_time_label.pack()

        # volume control
        # Creating scale scrollbars
        scale1 = tk.Scale(root, from_=0, to=100, orient=tk.HORIZONTAL, tickinterval=20, length=200, command=self.control_voice)
        scale1.pack()
        scale1.set(50)            # Set the default value of scale

        self.stop_button = tk.Button(bottom_frame, image=stop_icon,
                                     text=" pause", compound="left", command=self.toggle_pause,
                                     bg="#808C5C", fg="#FFFFFF")
        self.stop_button.image = stop_icon
        self.stop_button.pack(side=tk.LEFT, padx=10, pady=10, expand=True)

        next_button = tk.Button(bottom_frame, image=next_icon,
                                text=" next song", compound="left", command=self.next_song,
                                bg="#808C5C", fg="#FFFFFF")
        next_button.image = next_icon
        next_button.pack(side=tk.LEFT, padx=10, pady=10, expand=True)

        favorite_button = tk.Button(bottom_frame, image=favorite_icon,
                                    text=" favorite", compound="left",
                                    command=self.favorite_music, bg="#808C5C", fg="#FFFFFF")
        favorite_button.image = favorite_icon
        favorite_button.pack(side=tk.LEFT, padx=10, pady=10, expand=True)


        # lyrics button
        lyrics_button = tk.Button(bottom_frame, image=lyrics_icon,
                                    text=" Lyrics", compound="left",
                                    command=lambda: self.show_lyrics(self.selected_song_id), bg="#808C5C", fg="#FFFFFF")
        lyrics_button.image = lyrics_icon
        lyrics_button.pack(side=tk.LEFT, padx=10, pady=10, expand=True)

        #similar button
        recommend_button = tk.Button(bottom_frame, image=recommend_icon,
                                text=" play similar songs", compound="left", command=self.play_similar_song,
                                bg="#808C5C", fg="#FFFFFF")
        recommend_button.image = recommend_icon
        recommend_button.pack(side=tk.LEFT, padx=10, pady=10, expand=True)

    def update_treeview(self, music_list):
        # Update the tree view to show the current music list
        self.music_list = music_list
        self.tree.delete(*self.tree.get_children())
        for song in self.music_list:
            self.tree.insert('', 'end',
                             values=(song['id'], song['duration'], song['artist'], song['artist_id'], song['album']))

    def load_excel(self):
        # Load music list from Excel file
        global usingFileName 
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        usingFileName = filename
        if filename:
            self.music_list = self.load_music_from_excel(filename)
            self.update_treeview(self.music_list)

    def load_music_from_excel(self, filename):
        # Reading Excel files and loading music data
        wb = load_workbook(filename)
        ws = wb.active
        return [{'id': row[0], 'duration': row[1], 'artist': row[2], 'artist_id': row[3], 'album': row[4]}
                for row in ws.iter_rows(min_row=2, values_only=True)]

    def on_tree_select(self, event):
        # Get the selected song's ID
        selected_item = self.tree.focus()
        if selected_item:
            values = self.tree.item(selected_item, 'values')
            item = self.tree.selection()[0]
            self.current_song_index = self.tree.index(item)  # Stores the index of the currently selected song
            self.selected_song_id = values[0]
            self.selected_song_name = values[1]

    def get_song_url(self, song_id):
        url = f"https://dataiqs.com/api/netease/music/?type=songid&id={song_id}"
        response = requests.get(url)

        if response.status_code == 200:
            json_data = response.json()
            parsed_json = json.loads(json.dumps(json_data))
            song_url = parsed_json['song_url']
            return song_url
        else:
            return "Failed to retrieve data from the API"

    def search_music(self):
        # Implementation of the search function
        query = self.search_var.get()
        filtered_music_list = [song for song in self.music_list if query in str(song['duration']).lower()
                               or query in str(song['artist']).lower()
                               or query in str(song['artist_id'])
                               or query in str(song['album'])]
        self.update_treeview(filtered_music_list)

    def song_function(self):
        # Implementation of the hit song function
        plot_top_songs(root)

    def artist_function(self):
        # Implementation of the popular singers function
        # Creating a new window
        plot_window = tk.Toplevel(root)

        # Creating matplotlib graphs
        fig = Figure(figsize=(5, 4), dpi=100)
        plot = fig.add_subplot(111)

        # Data processing on DataFrame
        df = pd.read_excel('music_information.xlsx')
        df['Artist'] = df['Artist'].str.split(',').apply(pd.Series).stack().reset_index(drop=True)

        top_singer = df['Artist'].value_counts().nlargest(10)  # Get the top ten artists with the most appearances

        plt.rcParams['font.sans-serif'] = ['SimHei'] 
        
        # Plotting Horizontal Bars
        plot.barh(top_singer.index[::-1], top_singer.values[::-1])  # arrange in reverse order
        plot.set_xlabel('Appear time')
        plot.set_ylabel('Artist')
        plot.set_title('Most popular singers')

        # Embedding matplotlib graphs into the tkinter window
        canvas = FigureCanvasTkAgg(fig, master=plot_window)
        canvas.draw()
        canvas.get_tk_widget().pack()

    def analysis_function(self):
        # Implementation of data analysis functions
        # Assuming that songs are lists containing lyrics
        # Read all txt files in the lyrics folder
        hot_words={}

        lyrics_folder = 'lyrics'
        file_names = [f for f in os.listdir(lyrics_folder) if f.endswith('.txt')]
        
        stopwords_chinese = []
        # Open the deactivated words file
        with open('stopwords/哈工大停用词表.txt') as f:
            stopwords_chinese = f.read()
        chinese_stopwords_list = stopwords_chinese.split('\n')
        chinese_custom_stopwords_list = [i for i in chinese_stopwords_list]

        stopwords_english = []
        # Open the deactivated words file
        with open('stopwords/english.txt') as f:
            stopwords_english = f.read()
        english_stopwords_list = stopwords_english.split('\n')
        english_custom_stopwords_list = [i for i in english_stopwords_list]

        stopwords=[]
        stopwords.extend(chinese_custom_stopwords_list)
        stopwords.extend(english_custom_stopwords_list)

        # Create a new top-level window
        progress_window = tk.Toplevel()
        progress_window.title('Task progress')

        # Create a label to display progress information
        progress_label = tk.Label(progress_window, text='Task progress...')
        progress_label.pack()

        global analysis_stop_flag
        analysis_stop_flag = False
        def stop_task():
            # Discontinue a mission
            global analysis_stop_flag
            analysis_stop_flag=True
            # print("stop task")
            progress_window.destroy()
        stop_button = tk.Button(progress_window, text='Stop', command=stop_task)
        stop_button.pack()

        i=0
        progress_length=len(file_names)
        for file_name in file_names:
            if analysis_stop_flag:
                progress_label.config(text='Task stopped!')
                return
            file_path='lyrics/'+file_name

            progress_label.config(text=f'Working on track {i} song，{progress_length} total')
            progress_window.update()
            
            #print(i,'doing: '+str(file_path)+'\n')
            i=i+1
            with open(file_path, 'r', encoding='utf-8') as file:
                lines = file.readlines()  # Read lyrics line by line

            lyric = ""
            for line in lines:
                # Extract the content of each line of lyrics
                content = line.split(']')[-1]
                lyric += content.strip()  # Splice together the contents of the extracted lyrics
            word = jieba.lcut(lyric)  # decomposition word
            new_word = []
            for j in word:
                if len(j) >= 2:
                    new_word.append(j)  # Only add words with a length greater than 2
            final_text = " ".join(new_word)
            
            word_cloud = WordCloud(font_path="msyh.ttc",
                            width = 1000,
                            height = 700,
                            background_color='white',
                            max_words=100, stopwords=stopwords).generate(final_text)
            # Print the hot words for each song
            # Get word frequency data
            word_freq = word_cloud.process_text(final_text)

            # Write to dictionary
            top_words = sorted(word_freq, key=word_freq.get, reverse=True)[:5]
            # Update hot_words dictionary
            for word in top_words:
                if word in hot_words:
                    hot_words[word] += word_freq[word]
                else:
                    hot_words[word] = word_freq[word]
        
        progress_label.config(text='Task completed!')  # Mission Completion Tips
        progress_label.config(text='Task completed!')  # Mission Completion Tips
        progress_window.update()  # Update the progress tab immediately

        # After a delay of 1 second, close the window and execute the subsequent code.
        progress_window.after(1000, lambda: progress_window.destroy())

        #top_hot_words = dict(sorted(hot_words.items(), key=lambda item: item[1], reverse=True)[:20])
        # Print the final result
        #print(hot_words)

        # Converting a dictionary to a string
        hot_words_str = ' '.join(hot_words.keys())

        #geneartte word cloud
        word_cloud = WordCloud(font_path="msyh.ttc",
                        width = 1000,
                        height = 700,
                        background_color='white',
                        max_words=100, stopwords=stopwords).generate(hot_words_str)
            
        # Convert WordCloud objects to images
        wordcloud_image = word_cloud.to_image()

        # Creating a new matplotlib graph and displaying the image
        fig, ax = plt.subplots(figsize=(8, 6))
        ax.imshow(wordcloud_image, interpolation='bilinear')
        ax.axis('off')

        # Embedding matplotlib graphics into a Tkinter window
        # Creating a new window
        print('Start drawing the word cloud')
        plot_window = tk.Toplevel(root)
        canvas = FigureCanvasTkAgg(fig, master=plot_window)
        canvas.draw()
        canvas.get_tk_widget().pack()
        def close_window():
            plot_window.withdraw()
            
        plot_window.protocol("WM_DELETE_WINDOW", close_window)
        #save to dictionary
        output_file_name = 'word_cloud.png'  # Generated word cloud image file name
        word_cloud.to_file(output_file_name)  # Generate word clouds and save them as image files
        print("The name of the generated word cloud image file:", output_file_name)

    def play_music(self, song_id):
        # Implementation of the playback function
        # Implementation of the playback function
        if self.selected_song_id:
            try:
                song_url = self.get_song_url(song_id)
                response = requests.get(song_url, stream=True)
                if response.status_code == 200:
                    # Reads the response content into the BytesIO object
                    audio_stream = io.BytesIO(response.content)

                    #update length
                    audio_length = MP3(audio_stream).info.length
                    minutes, seconds = divmod(int(audio_length), 60)
                    total_time_str = f"total time:{minutes:02d}:{seconds:02d}"
                    self.total_time_label.config(text=total_time_str)  # Updating the Total Hours tab


                    # Passing audio streams to pygame mixer
                    
                    pygame.init()
                    # Setting the audio playback end event
                    pygame.mixer.music.set_endevent(pygame.constants.USEREVENT)
                    # Listen to the end of audio playback event
                    def poll():
                        for event in pygame.event.get():
                            if event.type == pygame.constants.USEREVENT:
                                self.play_similar_song()  # Audio playback ends and the next song is played
                        self.root.after(100, poll)  # Polling every 100 milliseconds

                    # Setting Polling Events
                    poll()

                    pygame.mixer.music.load(audio_stream, 'mp3')
                    pygame.mixer.music.play()
                    self.root.after(100, self.update_current_time)
                else:
                    messagebox.showerror("Error", f"Failed to get the song from URL: {song_url}")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def update_current_time(self):
        if pygame.mixer.music.get_busy():
            current_time = pygame.mixer.music.get_pos() / 1000  # Get the current playback time, converted to seconds
            minutes, seconds = divmod(int(current_time), 60)
            current_time_str = f"current time: {minutes:02d}:{seconds:02d}"
            self.current_time_label.config(text=current_time_str)  # Update current timestamp
            self.root.after(1000, self.update_current_time)  # Updates the current time every second

    def control_voice(self,value):
        voice = float(value) / 100
        pygame.mixer.music.set_volume(float(voice))
    
    def pause_music(self):
        # Implementation of the stop play function
        pygame.mixer.music.pause()

    def unpause_music(self):
        pygame.mixer.music.unpause()

    def toggle_pause(self):
        # Toggles the pause state and updates the button text
        global paused
        self.paused = not self.paused
        if self.paused:
            self.stop_button.config(image=self.play_icon, text=" resume")
            self.stop_button.image = self.play_icon
            # Perform a pause operation
            self.pause_music()
        else:
            self.stop_button.config(image=self.stop_icon, text=" pause")
            self.stop_button.image = self.stop_icon
            # Perform a resume pause operation
            self.unpause_music()

    def prev_song(self):
        # Implementation of the previous function
        if self.current_song_index is not None and len(self.music_list) > 0:
            if self.current_song_index > 0:

                current_index = self.current_song_index
                previous_index = current_index - 1
                if previous_index >= 0:
                    self.tree.selection_set(self.tree.get_children()[previous_index])
                    self.tree.focus(self.tree.get_children()[previous_index])
                    self.tree.see(self.tree.get_children()[previous_index])
                    self.current_song_index = previous_index

                previous_song_id = self.music_list[self.current_song_index]['id']
                self.selected_song_id=previous_song_id
                self.play_music(previous_song_id)  # Pass the ID of the last song
            else:
                messagebox.showinfo("Info", "This is the first song, can't play the previous one!")
        else:
            messagebox.showinfo("Info", "There are no songs in the list!")

    def play_similar_song(self):
        # Implement similar playback function
        if self.current_song_index is not None and len(self.music_list) > 0:
            if self.current_song_index >= 0 and self.current_song_index < len(self.music_list) - 1:

                current_index = self.current_song_index
                next_index = current_index

                self.tree.selection_set(self.tree.get_children()[next_index])
                self.tree.focus(self.tree.get_children()[next_index])
                self.tree.see(self.tree.get_children()[next_index])
                self.current_song_index = next_index

                stopwords_chinese = []
                # Open the deactivated words file
                with open('stopwords/哈工大停用词表.txt') as f:
                    stopwords_chinese = f.read()
                chinese_stopwords_list = stopwords_chinese.split('\n')
                chinese_custom_stopwords_list = [i for i in chinese_stopwords_list]

                stopwords_english = []
                # Open the deactivated words file
                with open('stopwords/english.txt') as f:
                    stopwords_english = f.read()
                english_stopwords_list = stopwords_english.split('\n')
                english_custom_stopwords_list = [i for i in english_stopwords_list]

                stopwords=[]
                stopwords.extend(chinese_custom_stopwords_list)
                stopwords.extend(english_custom_stopwords_list)
                
                #self.current_song_index += 1
                next_song_id = self.music_list[self.current_song_index]['id']
                current_song_id = next_song_id
                current_song_file_name='lyrics/'+str(current_song_id)+'.txt'

                #print('generating importwords for current song...')
                # Reading Lyrics Files
                with open(current_song_file_name, 'r', encoding='utf-8') as file:
                    lines = file.readlines()  # Read lyrics line by line
                lyric = ""
                for line in lines:
                    # Extract the content of each line of lyrics
                    content = line.split(']')[-1]
                    lyric += content.strip()  # Splice together the contents of the extracted lyrics

                # Process lyrics and generate word clouds
                #text = re.findall('[\u4e00-\u9fa5]+', lyric, re.S)  # Extract Chinese
                #text = " ".join(lyric)
                word = jieba.lcut(lyric)  # decomposition word
                new_word = [word for i in word if i not in stopwords]
                for i in word:
                    if len(i) >= 2:
                        new_word.append(i)  # Only add words with a length greater than 2
                current_text = (" ".join(word))

                #compare with all songs
                lyrics_folder = 'lyrics'
                file_names = [f for f in os.listdir(lyrics_folder) if f.endswith('.txt')]
                count_compare_time=0  
                max_similarity=0
                max_similarity_id=0
                for file_name in file_names:
                    if file_name.split('.')[0]==str(current_song_id):
                        continue
                    count_compare_time=count_compare_time+1
                    file_path='lyrics/'+file_name
                    with open(file_path, 'r', encoding='utf-8') as file:
                        lines = file.readlines()  # Read lyrics line by line
                    lyric = ""
                          
                    for line in lines:
                        # Extract the content of each line of lyrics
                        content = line.split(']')[-1]
                        lyric += content.strip()  # Splice together the contents of the extracted lyrics
                    word = jieba.lcut(lyric)  # decomposition word
                    selected_word = [word for i in word if i not in stopwords]
                    new_word = []
                    for i in selected_word:
                        if len(i) >= 2:
                            new_word.append(i)  # Only add words with a length greater than 2
                    final_text = (" ".join(word))
                    #print(final_text)
                    #print(type(final_text))
                    #compare with current song
                    similarity=compute_similarity(current_text,final_text)
                    if similarity>max_similarity:
                        # Generate a random number between 0 and 1
                        random_number = random.random()
                        if random_number<0.3:
                            max_similarity=similarity
                            max_similarity_id=file_name.split('.')[0]
                #play the similar song
                next_index=0
                self.play_music(max_similarity_id)  # Pass the ID of the next song
                for index, dictionary in enumerate(self.music_list):
                    if str(dictionary.get('id')) == max_similarity_id:
                        next_index=index
                        break
                #print('next index:',next_index)
                # update tree
                self.tree.selection_set(self.tree.get_children()[next_index])
                self.tree.focus(self.tree.get_children()[next_index])
                self.tree.see(self.tree.get_children()[next_index])
                self.current_song_index = next_index
                
                
    def next_song(self):
        # Implementation of the next song function
        if self.current_song_index is not None and len(self.music_list) > 0:
            if self.current_song_index >= 0 and self.current_song_index < len(self.music_list) - 1:

                current_index = self.current_song_index
                next_index = current_index + 1
                self.tree.selection_set(self.tree.get_children()[next_index])
                self.tree.focus(self.tree.get_children()[next_index])
                self.tree.see(self.tree.get_children()[next_index])
                self.current_song_index = next_index

                next_song_id = self.music_list[self.current_song_index]['id']
                self.selected_song_id=next_song_id
                self.play_music(next_song_id)  # Pass the ID of the next song

                self.tree.selection_set(self.tree.get_children()[next_index])
                self.tree.focus(self.tree.get_children()[next_index])
                self.tree.see(self.tree.get_children()[next_index])
                self.current_song_index = next_index
            else:
                messagebox.showinfo("Info", "This is the last song, can't play the next one!")
        else:
            messagebox.showinfo("Info", "There are no songs in the list!")

    def download_music(self):
        # Implementation of the download function
        # Folder selection dialogue box pops up
        folder_path = filedialog.askdirectory()
        if self.selected_song_id:
            song_url = self.get_song_url(self.selected_song_id)
            if not os.path.exists(folder_path):  # If the save directory does not exist, create it
                os.makedirs(folder_path)
            save_path = os.path.join(folder_path, '{}.mp3'.format(self.selected_song_name))  # Splice save path and file name
            if os.path.exists(save_path):  # Skip download if file already exists
                messagebox.showinfo("Download failed!", "{} already exists, skip download!".format(self.selected_song_name))
                return
            response = requests.get(song_url, headers=headers)  # Sending get requests with the requests module
            if response.status_code == 200:
                with open(save_path, 'wb') as f:  # Opening a file in binary write mode
                    f.write(response.content)
                messagebox.showinfo("Download successfully!", "{} download successfully!".format(self.selected_song_name))
            else:
                messagebox.showinfo("Download failed!", "{} download successfully!".format(self.selected_song_name))
        else:
            messagebox.showinfo("Info", "Please select the song to download!")

    def favorite_music(self):
        # Implementation of the collection function
        if self.selected_song_id:
            for song in self.music_list:
                if song['id'] == int(self.selected_song_id):
                    if song not in self.favorite_music_list:
                        self.favorite_music_list.append(song)
                        messagebox.showinfo("Info", f"{song['duration']} has been added to my favorite!")
                    else:
                        messagebox.showinfo("Info", "The song is already in my favorite!")
        else:
            messagebox.showinfo("Info", "Please select a song!")

    def show_favorite_song(self):
        self.update_treeview(self.favorite_music_list)

    def back_function(self):
        self.music_list = self.load_music_from_excel(usingFileName)
        self.update_treeview(self.music_list)

    # generate word cloud
    def generate_lyric_word_cloud(self, selected_song_id):
        selected_song_id = str(selected_song_id)
        file_name='lyrics/'+selected_song_id+'.txt'
            # Reading Lyrics Files
        with open(file_name, 'r', encoding='utf-8') as file:
            lines = file.readlines()  # Read lyrics line by line
        lyric = ""
        for line in lines:
            # Extract the content of each line of lyrics
            content = line.split(']')[-1]
            lyric += content.strip()  # Splice together the contents of the extracted lyrics
        # Process lyrics and generate word clouds
        word = jieba.lcut(lyric)  # decomposition word
        new_word = []
        for i in word:
            if len(i) >= 2:
                new_word.append(i)  # Only add words with a length greater than 2
        final_text = " ".join(word)
        #mask = np.array(Image.open("2.jpg"))  # Replace with user's image path
        stopwords = ["的","是","了","by","网易","网易云"] # Remove words that do not need to be displayed
        word_cloud = WordCloud(font_path="msyh.ttc",
                    width = 1000,
                    height = 700,
                    background_color='white',
                    max_words=100, stopwords=stopwords).generate(final_text)
            
        # Convert WordCloud objects to images
        wordcloud_image = word_cloud.to_image()
        # Create a new Toplevel window
        new_window = tk.Toplevel(root)

        # Convert images to formats supported by Tkinter
        tk_image = ImageTk.PhotoImage(wordcloud_image)

        # Create a Label and display the image
        label = tk.Label(new_window, image=tk_image)
        label.pack()

        new_window.mainloop()

    def show_lyrics(self, selected_song_id):
        print('loading lyrics...')
        filename = str(selected_song_id)+'.txt'
        filepath = os.path.join("lyrics", filename)
        if os.path.exists(filepath) and filepath.endswith(".txt"):
            with open(filepath, 'r', encoding='utf-8') as file:
                content = file.read()
                new_window = tk.Toplevel(root)
                new_window.title(filename)
                text_area = scrolledtext.ScrolledText(new_window, wrap=tk.WORD)
                text_area.insert(tk.INSERT, content)
                text_area.pack(expand=True, fill="both")
        else:
            messagebox.showerror("Error", "The file does not exist or is not a txt file")

if __name__ == '__main__':
    root = tk.Tk()
    app = MainUI(root)
    def on_closing():
    # Interrupt all processes when closing a window
        root.destroy()
    root.protocol("WM_DELETE_WINDOW", on_closing)
    root.mainloop()
